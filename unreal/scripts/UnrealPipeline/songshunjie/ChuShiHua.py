import unreal
import openpyxl as op


from Qt import QtCore
from Qt import QtWidgets


from dayu_widgets.field_mixin import MFieldMixin
from dayu_widgets.push_button import MPushButton
from dayu_widgets.line_edit import MLineEdit
from dayu_widgets import dayu_theme
from dayu_widgets.qt import application

print('ChuShiHua')


class mw(QtWidgets.QWidget, MFieldMixin):

    sc=[]
    cam=[]
    startK=[]
    endK=[]
    cam_se=[]
    ep=''
    #文件类型
    flie_class=['Lighting','Animation','Cache','VFX','Modify']  
    #文件路径
    excel_path=''

    def __init__(self, parent=None):
        super().__init__(parent)
        
        self.uii()

    def uii(self):   
        self.setWindowTitle('初始化项目基础文件目录')
        self.resize(300,200)
        lay=QtWidgets.QVBoxLayout()

        folder_lay=QtWidgets.QVBoxLayout()

        #导入文件
        create_base_folder=MPushButton(text="创建基础文件夹")
        create_base_folder.clicked.connect(self.baseDirectory)
        self.flie_import=MLineEdit().file().medium()
        self.flie_import.setPlaceholderText(self.tr("选择需要导入信息的Excel文件"))
        create_folder=MPushButton(text="创建shots文件夹")
        create_folder.clicked.connect(self.createDirectory)

        folder_lay.addWidget(create_base_folder)
        folder_lay.addWidget(self.flie_import)
        folder_lay.addWidget(create_folder)

        import_lay=QtWidgets.QVBoxLayout()
        




        lay.addLayout(folder_lay)
        lay.addLayout(import_lay)
        self.setLayout(lay)

    def directory_list(self):

        self.excel_path=self.flie_import.text()

        df=op.load_workbook(self.excel_path,data_only=True)
        sheet=df.get_sheet_by_name('镜头表')
        self.sc=[]
        self.cam=[]
        self.cam_se=[]
        self.startK=[]
        self.endK=[]

        all_list=[]
        i=0
        #确定行和列
        rowcount=sheet.max_row
        colcount=8
        #读取所需的Excel内容
        for i in range(13,rowcount+1):
            list=[]
            for j in range(3,colcount+1):
                list.append(sheet.cell(row=i,column=j).value)
            all_list.append(list)
        # print(all_list)

        #数据分配
        
        ep=all_list[0][0]
        for i in all_list:
            cam_s=[]
            if i[0]==ep and i[1] and i[2] and i[3] and i[4] and i[5]:
                self.sc.append(i[1])
                self.cam.append(i[2])
                # self.startK.append(i[3])
                # self.endK.append(i[4])
                cam_s.append(i[2])
                cam_s.append(i[3])
                cam_s.append(i[4])
                self.cam_se.append(cam_s)

        return ep


    def createDirectory(self):

        # print(self.sc)
        # print(len(self.sc))

        ep=self.directory_list()
        sc_list=[]

        #简化sc数据内容
        for ii in self.sc:
            if ii in sc_list:
                pass
            else:
                sc_list.append(ii)
        # print(sc_list)

        #对镜头信息分组
        cam_dict={}

        for sc_list_i in sc_list:
            cam_list=[]
            for cam_i in self.cam_se:
                if cam_i[0].split('_',2)[1]==sc_list_i:
                    cam_list.append(cam_i)
            #排序列表
            cam_list.sort()
            print(cam_list)
            cam_dict['%s'%sc_list_i]=cam_list
        
       
        lt_ls_list=[]
        an_ls_list=[]
        # 创建文件夹和所需的文件
        unreal.EditorAssetLibrary.make_directory('/Game/Shots/%s/Preview'%(ep))
        for sc_name in sc_list:

            if not unreal.EditorAssetSubsystem().does_asset_exist('/Game/Shots/%s/Preview/%s_%s_Pv_Map'%(ep,ep,sc_name)):
                unreal.AssetToolsHelpers.get_asset_tools().create_asset(asset_name='%s_%s_Pv_Map'%(ep,sc_name),package_path='/Game/Shots/%s/Preview'%(ep),asset_class=unreal.World,factory=unreal.WorldFactory())

            if not unreal.EditorAssetSubsystem().does_asset_exist('/Game/Shots/%s/Preview/%s_%s_Preview'%(ep,ep,sc_name)):
                preview_sequence=unreal.AssetToolsHelpers.get_asset_tools().create_asset(asset_name='%s_%s_Preview'%(ep,sc_name),package_path='/Game/Shots/%s/Preview'%(ep),asset_class=unreal.LevelSequence,factory=unreal.LevelSequenceFactoryNew())
                preview_sequence:unreal.LevelSequence
                preview_sequence.set_display_rate((25,1))
                preview_track=preview_sequence.add_track(unreal.MovieSceneSubTrack)
                preview_track:unreal.MovieSceneSubTrack
                preview_track.set_display_name('%s_%s'%(ep,sc_name))

            start_key = 0
            end_key = 0
                
            unreal.EditorAssetLibrary.save_directory('/Game/Shots/%s/Preview'%(ep))
            unreal.LevelEditorSubsystem().load_level('/Game/Shots/%s/Preview/%s_%s_Pv_Map'%(ep,ep,sc_name))

            for cam_name in cam_dict[sc_name]:
                flie_name=cam_name[0]
                for file_class_name in self.flie_class:
                    if file_class_name=='Lighting':
                        # flie_name='%s_lt'%(cam_name[0])
                        if not unreal.EditorAssetSubsystem().does_asset_exist('/Game/Shots/%s/%s/%s/%s/%s_lt'%(ep,sc_name,flie_name,file_class_name,cam_name[0])):
                            unreal.AssetToolsHelpers.get_asset_tools().create_asset(asset_name='%s_lt'%(cam_name[0]),package_path='/Game/Shots/%s/%s/%s/%s'%(ep,sc_name,flie_name,file_class_name),asset_class=unreal.World,factory=unreal.WorldFactory())
                        if not unreal.EditorAssetSubsystem().does_asset_exist('/Game/Shots/%s/%s/%s/%s_Render'%(ep,sc_name,flie_name,flie_name)):
                            lt_ls=unreal.AssetToolsHelpers.get_asset_tools().create_asset(asset_name='%s_Render'%(flie_name),package_path='/Game/Shots/%s/%s/%s'%(ep,sc_name,flie_name),asset_class=unreal.LevelSequence,factory=unreal.LevelSequenceFactoryNew())
                            lt_ls:unreal.LevelSequence
                            lt_ls.set_display_rate((25,1))
                            if cam_name[1]:
                                lt_ls.set_playback_start(int(cam_name[1])-1)
                            if cam_name[2]:
                                lt_ls.set_playback_end(int(cam_name[2])+1)
                            lt_ls_list.append(lt_ls)
                        unreal.EditorAssetLibrary.save_directory('/Game/Shots/%s/%s/%s/%s'%(ep,sc_name,flie_name,file_class_name))
                        unreal.LevelEditorSubsystem().load_level('/Game/Shots/%s/%s/%s/%s/%s_lt'%(ep,sc_name,flie_name,file_class_name,cam_name[0]))

                        #将灯光序列放入总序列
                        end_key += lt_ls.get_playback_end()
                        
                        # preview_track.set_display_name(lt_ls.get_name())
                        preview_section=preview_track.add_section()
                        preview_section.set_sequence(lt_ls)
                        preview_section.set_range(start_key,end_key)
                        preview_section.set_row_index(0)
                        
                        preview_sequence.set_view_range_end(end_key*1.03/25)
                        preview_sequence.set_playback_end(end_key)
                        # preview_sequence.set_work_range_end(end_key)
                        
                        start_key += lt_ls.get_playback_end()
                    
                    if file_class_name=='Animation':
                        # flie_name='%s_an'%(cam_name[0])
                        if not unreal.EditorAssetSubsystem().does_asset_exist('/Game/Shots/%s/%s/%s/%s/%s'%(ep,sc_name,flie_name,file_class_name,flie_name)):
                            an_ls=unreal.AssetToolsHelpers.get_asset_tools().create_asset(asset_name='%s_an'%(flie_name),package_path='/Game/Shots/%s/%s/%s/%s'%(ep,sc_name,flie_name,file_class_name),asset_class=unreal.LevelSequence,factory=unreal.LevelSequenceFactoryNew())
                            an_ls.set_display_rate((25,1))
                            if cam_name[1]:
                                an_ls.set_playback_start(int(cam_name[1])-1)
                            if cam_name[2]:
                                an_ls.set_playback_end(int(cam_name[2])+1)
                            an_ls_list.append(an_ls)
                    if file_class_name=='Cache':
                        # flie_name='%s_cache'%(cam_name[0])
                        unreal.EditorAssetLibrary.make_directory('/Game/Shots/%s/%s/%s/%s'%(ep,sc_name,flie_name,file_class_name))
                    if file_class_name=='VFX':
                        # flie_name='%s_vfx'%(cam_name[0])
                        unreal.EditorAssetLibrary.make_directory('/Game/Shots/%s/%s/%s/%s/VFX_DT'%(ep,sc_name,flie_name,file_class_name,))
                    if file_class_name=='Modify':
                        # flie_name='%s_modify'%(cam_name[0])
                        unreal.EditorAssetLibrary.make_directory('/Game/Shots/%s/%s/%s/%s'%(ep,sc_name,flie_name,file_class_name))

                #创建总关卡
                if not unreal.EditorAssetSubsystem().does_asset_exist('/Game/Shots/%s/%s/%s/%s_Map'%(ep,sc_name,flie_name,flie_name)):
                    unreal.AssetToolsHelpers.get_asset_tools().create_asset(asset_name='%s_Map'%(flie_name),package_path='/Game/Shots/%s/%s/%s'%(ep,sc_name,flie_name),asset_class=unreal.World,factory=unreal.WorldFactory())

            
                    # lt_ls.set_display_rate((25,1))
                    # if cam_name[1]:
                    #     lt_ls.set_playback_start(int(cam_name[1])-1)
                    # if cam_name[2]:
                    #     lt_ls.set_playback_end(int(cam_name[2])+1)
                    # lt_ls_list.append(lt_ls)
                
                # if not unreal.EditorAssetSubsystem().does_asset_exist('/Game/Shots/%s/%s/%s/%s_Render'%(ep,sc_name,flie_name,flie_name)):
                #     lt_ls=unreal.AssetToolsHelpers.get_asset_tools().create_asset(asset_name='%s_Render'%(flie_name),package_path='/Game/Shots/%s/%s/%s'%(ep,sc_name,flie_name),asset_class=unreal.LevelSequence,factory=unreal.LevelSequenceFactoryNew())
        # unreal.EditorAssetLibrary.make_directory('/Game/Shots/%s/Lighting/KeyLight'%(ep))
        # unreal.EditorAssetLibrary.make_directory('/Game/Shots/%s/VFX/VFX_Cache'%(ep))
        

            
        for lt in lt_ls_list:
            for an in an_ls_list:
                if lt.get_name().rsplit('_',1)[0]==an.get_name().rsplit('_',1)[0]:
                    # lt.add_spawnable_from_instance(an) 
                    lt_track=lt.add_track(unreal.MovieSceneSubTrack)
                    lt_section=lt_track.add_section()
                    lt_section.set_sequence(an)
                    lt_section.set_range(an.get_playback_start(),an.get_playback_end())
                    break
        #保存全部创建的文件
        unreal.EditorAssetLibrary.save_directory('/Game/Shots')

    def baseDirectory(self):
        #AAI文件夹
        unreal.EditorAssetLibrary.make_directory('/Game/AAI/Reference/Character')
        unreal.EditorAssetLibrary.make_directory('/Game/AAI/Reference/Pro')
        unreal.EditorAssetLibrary.make_directory('/Game/AAI/Reference/Scenes')
        unreal.EditorAssetLibrary.make_directory('/Game/AAI/Reference/Common')

        #Assets文件夹
        unreal.EditorAssetLibrary.make_directory('/Game/Assets/Common')
        unreal.EditorAssetLibrary.make_directory('/Game/Assets/Character')
        unreal.EditorAssetLibrary.make_directory('/Game/Assets/Pro')
        unreal.EditorAssetLibrary.make_directory('/Game/Assets/Scenes')

        #Shots文件夹
        unreal.EditorAssetLibrary.make_directory('/Game/Shots/Lighting')


        #保存全部创建的文件
        unreal.EditorAssetLibrary.save_directory('/Game')

def start():
    with application() as app:
        global test
        test = mw()
        dayu_theme.apply(test)
        test.show()
        unreal.parent_external_window_to_slate(int(test.winId()))



if __name__ == "__main__":

    start()
