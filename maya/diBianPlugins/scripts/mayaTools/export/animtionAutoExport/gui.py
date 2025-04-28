#coding=utf-8

from maya import cmds,mel
import openpyxl as op
import time
import os


def ExportFbx(file_path):
    if not cmds.pluginInfo('fbxmaya', q=True, loaded=True):
        pm.loadPlugin("fbxmaya")

    mel.eval("FBXProperty Export|IncludeGrp|Geometry|BlindData -v false")

    fbx_export_options = [
        ("-s",),  # Selection Only
        ("-v", "FBX201400"),  # FBX Version
        ("-f", "MotionBuilder.fbx"),  # File Name
        # ("-es",),  # Embed Textures
        # ("-ea",),  # Embed Media
        ("-fr", "25"),  # Frame Rate
        ("-q",),  # Quiet Mode
        ("-p", "256"),  # Optimization Level
        # ("-axisUp", "y"),  # Up Axis
        # ("-axisFront", "z"),  # Front Axis
        ("-a", "model", "camera"),  # Animation Only (Model and Camera)
        ("-enablePointCache",),  # Point Cache
        ("-enableSampling",),  # Animation Sampling
        ("-stripNamespaces",),  # Strip Namespace
        ("-filterType", "AnimCurve"),  # Filter Type
        # ("-fcp", "0"),  # Bake Start Time
        # ("-fcp", "0"),  # Bake End Time
        ("-tp", "0"),  # Use Default Take
        ("-simplify", "1"),  # Simplify Animation Curves
    ]

    cmds.file(file_path, force=True, options=";".join([str(opt) for opt in fbx_export_options]), type="FBX export", exportSelected=True)

class win():
    def __init__(self):
        self.winName=u"自动导出动画FBX文件"
        if cmds.window(self.winName,q=1,ex=1):
            cmds.deleteUI(self.winName)
        cmds.window(self.winName,widthHeight=(400,260))
        self.UI1()
        
        
    def UI1(self):
        self.start=1
        self.end=5
        self.column1=cmds.columnLayout( adjustableColumn=True )
        
        #cmds.textFieldGrp('start',label='设置烘焙起始帧',text='%s'%self.start,cal=(1,'left'),cw2=(100,50))
        #cmds.textFieldGrp('end',label='设置烘焙结束帧',text='%s'%self.end,cal=(1,'left'),cw2=(100,50))
        
        
                
        cmds.text(u'选择Excel文件',align='left')
        cmds.rowLayout( numberOfColumns=2, columnWidth2=(80, 75))
        cmds.textField('excel',w=370)
        cmds.symbolButton( image='circle.png',w=20,i="navButtonBrowse.xpm",c=self.openExcel)
        cmds.setParent(self.column1)
        
        
        
        cmds.text(u'选择要执行的文件夹',align='left')
        cmds.rowLayout( numberOfColumns=2, columnWidth2=(80, 75))
        cmds.textField('export',w=370)
        cmds.symbolButton( image='circle.png',w=20,i="navButtonBrowse.xpm",c=self.openFile)
        cmds.setParent(self.column1)
        
        cmds.button(label=u'检测Excel文件与所选文件夹是否匹配',command=self.detectionFile)
        cmds.setParent(self.column1)
        cmds.text('detection',label='',align='left',ebg=0,bgc=[1,0,0])
        
        #预留帧
        cmds.rowLayout( numberOfColumns=2)
        cmds.textFieldGrp('start_offset',label=u'设置整体偏移帧',text=10,cal=(1,'left'),cw2=(100,50))
        cmds.textFieldGrp('end_offset',label=u'设置预留结束帧',text=5,cal=(1,'left'),cw2=(100,50))
        cmds.setParent(self.column1)
        
        cmds.rowLayout( numberOfColumns=2, columnWidth2=(80, 75))
        cmds.textField('fbx_path',text='',w=370)
        cmds.symbolButton( image='circle.png',w=20,i="navButtonBrowse.xpm",c=self.exportFile)
        cmds.setParent(self.column1)
        
        cmds.button(label=u'烘焙并导出FBX到指定位置',command=self.execute)
        cmds.setParent(self.column1)
        
        cmds.text(u'执行前请确保已经获取所有所需的引用文件',align='left')
        cmds.setParent(self.column1)
        
        
    def excelGet(self,*args):
        
        #获取Excel表格信息
        excel_path=cmds.textField('excel',text=1,q=1)
        df=op.load_workbook(excel_path,data_only=True)
        sheet=df.get_sheet_by_name(u'镜头表')
        
        sc=[]
        cam=[]
        startK=[]
        endK=[]
        start_endK=[]
        all_list=[]
        i=0
        #确定行和列
        rowcount=sheet.max_row
        colcount=7
        #读取所需的Excel内容
        for i in range(13,rowcount+1):
            list=[]
            for j in range(3,colcount+1):
                list.append(sheet.cell(row=i,column=j).value)
            all_list.append(list)
        
        ep=all_list[0][0]
        self.start_endK_d={}
        for i in all_list:
            if i[0]==ep:
                start_end_c=[]
                sc.append(i[1])
                cam.append(i[2])
                start_end_c.append(i[3])
                start_end_c.append(i[4])
                start_endK.append(start_end_c)
                self.start_endK_d[i[2]]=start_end_c    #key为场次名称 value为起始结束帧
        # print(self.start_endK_d)
        
            
    def execute(self,*args):
        
        export_fail_list=[]
        
        self.excelGet()
        #获取文件路径
        dir_path =cmds.textField('export',text=1,q=1)
        ma_path=[]
        for dirpath, dirnames, filenames in os.walk(dir_path):
            for filename in filenames:
                if '.ma' in filename or '.mb' in filename:
                    ma_path.append(os.path.join(dirpath, filename))
                    
        if ma_path:
            #通过文件路径打开ma文件
            for ma in ma_path:
                #print(cmds.file(q=1,sn=1))
                cmds.file(modified=0)
                cmds.file(ma,o=1)
                time.sleep(5)
                #将帧速率设置为25
                cmds.currentUnit( time='pal' )
                #获取当前场景的场次信息
                key=cmds.file(q=1,sn=1).rsplit('/',1)[-1].rsplit('_',1)[0]
                
                if key in self.start_endK_d:
                    self.start=self.start_endK_d[key][0]#获取关键帧起始信息
                    self.end=self.start_endK_d[key][1]
                    
                    #获取偏移值
                    start_offset=int(cmds.textFieldGrp('start_offset',text=1,q=1))
                    end_offset=int(cmds.textFieldGrp('end_offset',text=1,q=1))
                    
                    start_key=self.start
                    end_key=self.end+start_offset+end_offset
                    
                    #设置时间轴范围    
                    cmds.playbackOptions( minTime=start_key, maxTime=end_key )

                    #获取全部名称空间名
                    space_names=cmds.namespaceInfo( lon=True )
                    for self.space_name in space_names:
                        print('')
                        if cmds.objExists(self.space_name+':Root_M'):
                            #选择所有对应名称空间的对象
                            #cmds.select(self.space_name+':*',hi=1)
                            #将所有对象放入列表
                            #frame_list=cmds.ls(sl=1)
                            #偏移关键帧
                            #start_offset=int(cmds.textFieldGrp('start_offset',text=1,q=1))
                            #cmds.keyframe( frame_list,relative=True,timeChange=start_offset)
                            
                            print('ch',self.space_name)
                            cmds.select(self.space_name+':DeformationSystem')
                            #cmds.pickWalk(self.space_name+':DeformationSystem', direction='down')
                            
                            try:
                                #cmds.select(self.space_name+':Geometry',add=1)
                                self.bake()
                                self.exportFBX()
                                print('ch',self.space_name)
                            except:
                                export_fail_list.append(ma+'  :  '+self.space_name)
                                print('!!!!!!!!!!!!'+self.space_name+' export fail')
                            
                        elif cmds.objExists(self.space_name+':*_GuGe_G'):

                            print('pro',self.space_name)
                            cmds.select(self.space_name+':*_GuGe_G')
                            
                            try:
                                # cmds.select(self.space_name+':*_Pro_Mo',add=1)
                                self.bake()
                                self.exportFBX()
                                print('pro',self.space_name)
                            except:
                                export_fail_list.append(ma+'  :  '+self.space_name)
                                print('!!!!!!!!!!!!'+self.space_name+' export fail')
                else :
                    print(cmds.file(q=1,sn=1)+'  与表内信息不匹配')
                    # cmds.text('detection',label=u'Excel数据与当前场景不匹配',ebg=1,e=1)
                    # break
            for export_fail in export_fail_list:
                print(export_fail)
    def bake(self,*args):
        
        obj=cmds.ls(sl=1)

        #导入选定对象引用并删除选定对象的名称空间
        path_reference=cmds.referenceQuery(obj[0], f=True )
        cmds.file(path_reference, ir=1)
        # cmds.parent(self.obj,world=True)#解除组
        self.obj_namespace=str(obj[0]).split(':',1)[0]+':'
        cmds.select(obj)
        cmds.namespace(removeNamespace = self.obj_namespace, mergeNamespaceWithParent = True)
        self.obj_new=cmds.ls(sl=1)
        
        #选择需要烘焙的对象
        cmds.select(self.obj_new,hi=1)
        # cmds.select(cmds.ls(type= "blendShape"),add=1)
        
        #烘焙
        self.obj_bake_l=[]
        self.obj_list=cmds.ls(sl=1)
        i=0
        while i<len(self.obj_list):
            self.obj_bake_l.append(str(self.obj_list[i]))
            i+=1
        self.obj_bake_a=str(self.obj_bake_l)[2:-2]
        self.obj_bake=self.obj_bake_a.replace("'",'"')
        #self.start=int(cmds.textFieldGrp('start',text=1,q=1))
        #self.end=int(cmds.textFieldGrp('end',text=1,q=1))
        
        

        
        #获取偏移值
        start_offset=int(cmds.textFieldGrp('start_offset',text=1,q=1))
        end_offset=int(cmds.textFieldGrp('end_offset',text=1,q=1))
        
        
        #设置烘焙起始结束帧
        self.start=self.start
        self.end=self.end
        print(self.start,self.end)
        mel.eval('''bakeResults -simulation true -t "%s:%s" -sampleBy 1 -oversamplingRate 1 -disableImplicitControl true -preserveOutsideKeys true -sparseAnimCurveBake false -removeBakedAttributeFromLayer false -removeBakedAnimFromLayer false -bakeOnOverrideLayer false -minimizeRotation true -controlPoints false -shape true {"%s"}'''%(self.start,self.end,self.obj_bake))
        
        
        #选择所有烘焙的对象
        cmds.select(self.obj_new,hi=1)
        #将所有对象放入列表
        frame_list=cmds.ls(sl=1)
        #偏移关键帧
        cmds.keyframe( frame_list,relative=True,timeChange=start_offset)
        
        #烘焙前后帧
        #cmds.bakeResults( frame_list, t=(self.start,self.end+end_offset+start_offset), simulation=True )
        if start_offset!=0:
            cmds.bakeResults( frame_list, t=(self.start,self.start+start_offset+1), simulation=True,preserveOutsideKeys=True )
        if end_offset!=0:
            cmds.bakeResults( frame_list, t=(self.end,self.end+start_offset+end_offset), simulation=True,preserveOutsideKeys=True )
        
        
        
        #删除多余节点
        #cmds.select( '*lower*ipBC*','*upper*ipBC*','*lower*ipBC*_*' )
        #cmds.delete()
        
        #选择所有需要导出的对象
        cmds.select(self.obj_new,hi=1)
    
    
        
    def exportFBX(self,*args):
        
        #设置FBX文件名
        #text_path=str(cmds.file(q=1,sn=1)).rsplit('_',1)[0]+'_FBX'
        #if 'Mb' in text_path:
        #    text_path_old=text_path.replace('Mb','FBX',1)
        #elif 'Lighting/File' in text_path:
        #    text_path_old=text_path.replace('Lighting/File','Animation/FBX',1)#创建路径
        #else :
        #    text_path_old=text_path
        
        
        
        #获取文件路径并创建新路径
        dir_path = cmds.textField('fbx_path',text=1,q=1)
        sc_folder = str(cmds.file(q=1,sn=1)).rsplit('_',1)[0].rsplit('/',1)[-1]
        sc_name = sc_folder.split('_')[1]
        text_path_old=dir_path+'/'+sc_name+'/'+sc_folder
        save_file_name='/'+str(cmds.file(q=1,sn=1)).rsplit('_',1)[0].rsplit('/',1)[-1]+'_an_'+self.space_name#创建文件名
        self.export_path=text_path_old+save_file_name
        print(self.export_path)
        #导出fbx
        cmds.select(self.obj_new)
        cmds.FBXResetExport()
        mel.eval('FBXExportFileVersion -v FBX201300')
        mel.eval('FBXExportSmoothingGroups -v true')
        mel.eval('FBXExportUpAxis y')
        mel.eval('FBXExportSmoothMesh -v true')
        start_offset=int(cmds.textFieldGrp('start_offset',text=1,q=1))
        end_offset=int(cmds.textFieldGrp('end_offset',text=1,q=1))
        start_key=self.start
        end_key=self.end+start_offset+end_offset
        #mel.eval('FBXExportSplitAnimationIntoTakes -v \"Take_001 " %s %s'%(self.start,self.end+start_offset+end_offset))
        #mel.eval('FBXExportSplitAnimationIntoTakes -c ')
        
        
        
        #创建对应FBX文件夹
        if not os.path.exists(self.export_path.rsplit('/',1)[0]):
            os.makedirs(self.export_path.rsplit('/',1)[0])
        
        if os.path.exists(self.export_path+'.fbx'):
            os.remove(self.export_path+'.fbx')
            print('delete '+self.export_path+'.fbx')
        
        #设置时间轴范围    
        cmds.playbackOptions( minTime=start_key, maxTime=end_key )
        #导出
        mel.eval('FBXExport -f "%s" -s'%(self.export_path))
        
        #eval('file -force -options "" -typ "FBX export" -pr -es "%s.fbx"'%(self.export_path))
        
        #ExportFbx(self.export_path)
        
        #创建收纳组
        collect_g='%s_g'%self.obj_namespace.split(':',1)[0]
        cmds.group(em=True,n=collect_g)
        cmds.parent(self.obj_new,collect_g)

    def openFile(self,*args):
        #提示栏清空
        cmds.text('detection',label='',align='left',ebg=0,bgc=[1,0,0],e=1)
        
        singleFilter=u"目录"
        self.open_path=str(cmds.fileDialog2(fileFilter=singleFilter,startingDirectory =str(cmds.file(q=1,sn=1)).rsplit('/',1)[0],fm=3)[0])
        cmds.textField('export',text=self.open_path,e=1)
        
    def exportFile(self,*args):
        #提示栏清空
        cmds.text('detection',label='',align='left',ebg=0,bgc=[1,0,0],e=1)
        
        singleFilter=u"目录"
        self.open_path=str(cmds.fileDialog2(fileFilter=singleFilter,startingDirectory =str(cmds.file(q=1,sn=1)).rsplit('/',1)[0],fm=3)[0])
        cmds.textField('fbx_path',text=self.open_path,e=1)
        
    def openExcel(self,*args):
        #提示栏清空
        cmds.text('detection',label='',align='left',ebg=0,bgc=[1,0,0],e=1)
        
        singleFilter="*.xlsx"
        self.open_path=str(cmds.fileDialog2(fileFilter=singleFilter,startingDirectory=str(cmds.file(q=1,sn=1)).rsplit('/',1)[0],fileMode=4)[0])
        cmds.textField('excel',text=self.open_path,e=1)
        
    
    
    def detectionFile(self,*args):
        
        self.excelGet()
        #获取文件路径
        dir_path =cmds.textField('export',text=1,q=1)
        ma_file=[]
        for dirpath, dirnames, filenames in os.walk(dir_path):
            for filename in filenames:
                if '.ma' in filename or '.mb' in filename:
                    ma_file.append(filename.rsplit('_',1)[0])
        ma_exist=False
        error_exist=False
        for ma_name in ma_file:    
            
            if ma_name not in self.start_endK_d:
                # cmds.text('detection',label=u'Excel与当前所选文件夹内容不匹配',ebg=1,bgc=[1,0,0],e=1)
                print(ma_name)
                error_exist=True
            else:
                ma_exist=True
                # cmds.text('detection',label=u'Excel与当前所选文件夹内容匹配',ebg=0,bgc=[0,0,0],e=1)
            
        if error_exist and not ma_exist:
            cmds.text('detection',label=u'Excel与当前所选文件夹内容不匹配',ebg=1,bgc=[1,0,0],e=1)
        elif error_exist and ma_exist:
            cmds.text('detection',label=u'Excel与当前所选文件夹内容有部分不匹配',ebg=1,bgc=[1,1,0],e=1)
        else:
            cmds.text('detection',label=u'Excel与当前所选文件夹内容匹配',ebg=0,bgc=[0,0,0],e=1)
                
                

def showUI():
    win()
    cmds.showWindow()
    
    
    
if __name__=='__main__':
    showUI()
 


