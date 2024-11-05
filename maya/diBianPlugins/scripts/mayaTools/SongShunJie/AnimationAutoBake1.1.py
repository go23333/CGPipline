# -*- coding: utf-8 -*-

from maya import cmds,mel
import openpyxl as op
import time
import os

class win():
	def __init__(self):
		self.winName="�Զ���������FBX�ļ�"
		if cmds.window(self.winName,q=1,ex=1):
			cmds.deleteUI(self.winName)
		cmds.window(self.winName,widthHeight=(400,200))
		self.UI1()
		
		
	def UI1(self):
		self.start=1
		self.end=5
		self.column1=cmds.columnLayout( adjustableColumn=True )
		
		#cmds.textFieldGrp('start',label='���ú決��ʼ֡',text='%s'%self.start,cal=(1,'left'),cw2=(100,50))
		#cmds.textFieldGrp('end',label='���ú決����֡',text='%s'%self.end,cal=(1,'left'),cw2=(100,50))
		
		cmds.setParent(self.column1)
		
		cmds.text('ѡ��Excel�ļ�',align='left')
		cmds.rowLayout( numberOfColumns=2, columnWidth2=(80, 75))
		cmds.textField('excel',w=370)
		cmds.symbolButton( image='circle.png',w=20,i="navButtonBrowse.xpm",c=self.openExcel)
		cmds.setParent(self.column1)
		
		
		
		cmds.text('ѡ��Ҫִ�е��ļ���',align='left')
		cmds.rowLayout( numberOfColumns=2, columnWidth2=(80, 75))
		cmds.textField('export',w=370)
		cmds.symbolButton( image='circle.png',w=20,i="navButtonBrowse.xpm",c=self.openFile)
		cmds.setParent(self.column1)
		
		cmds.button(label='���Excel�ļ�����ѡ�ļ����Ƿ�ƥ��',command=self.detectionFile)
		cmds.setParent(self.column1)
		cmds.text('detection',label='',align='left',ebg=0,bgc=[1,0,0])
		
		cmds.button(label='�決������FBX��ָ��λ��',command=self.execute)
		cmds.setParent(self.column1)
		
		cmds.text('ִ��ǰ��ȷ���Ѿ���ȡ��������������ļ�',align='left')
		cmds.setParent(self.column1)
		
		
	def excelGet(self,*args):
		
		#��ȡExcel�����Ϣ
		excel_path=cmds.textField('excel',text=1,q=1)
		df=op.load_workbook(excel_path)
		sheet=df.get_sheet_by_name('��ͷ��')
		
		sc=[]
		cam=[]
		startK=[]
		endK=[]
		start_endK=[]
		all_list=[]
		i=0
		#ȷ���к���
		rowcount=sheet.max_row
		colcount=7
		#��ȡ�����Excel����
		for i in range(13,rowcount+1):
			list=[]
			for j in range(3,colcount+1):
			    list.append(sheet.cell(row=i,column=j).value)
			all_list.append(list)
		
		ep=all_list[0][0]
		self.start_endK_d={}
		for i in all_list:
		    if i[0]==ep:
		    	start_end_c=[]
		        sc.append(i[1])
		        cam.append(i[2])
		        start_end_c.append(i[3])
		        start_end_c.append(i[4])
		        start_endK.append(start_end_c)
		        self.start_endK_d[i[2]]=start_end_c	#keyΪ�������� valueΪ��ʼ����֡
		#print(self.start_endK_d)
		
			
	def execute(self,*args):
		
		self.excelGet()
		#��ȡ�ļ�·��
		dir_path =cmds.textField('export',text=1,q=1)
		ma_path=[]
		for dirpath, dirnames, filenames in os.walk(dir_path):
			for filename in filenames:
				if '.ma' in filename or '.mb' in filename:
					ma_path.append(os.path.join(dirpath, filename))
					
		if ma_path:
			#ͨ���ļ�·����ma�ļ�			
			for ma in ma_path:
				cmds.file(modified=0)
				cmds.file(ma,o=1)
				time.sleep(5)
				#��ȡ��ǰ�����ĳ�����Ϣ
				key=cmds.file(q=1,sn=1).rsplit('/',1)[-1].rsplit('_',1)[0]
				if key in self.start_endK_d:
					self.start=self.start_endK_d[key][0]#��ȡ�ؼ�֡��ʼ��Ϣ
					self.end=self.start_endK_d[key][1]

					#��ȡȫ�����ƿռ���
					space_names=cmds.namespaceInfo( lon=True )
					for self.space_name in space_names:
						if cmds.objExists(self.space_name+':DeformationSystem'):
							cmds.select(self.space_name+':DeformationSystem')
							#cmds.pickWalk(self.space_name+':DeformationSystem', direction='down' )
							#cmds.select(self.space_name+':Geometry',add=1)
							self.bake()
							self.exportFBX()
						elif cmds.objExists(self.space_name+':DaoJu_GuGe_G*'):
							cmds.select(self.space_name+':DaoJu_GuGe_G*')
							self.bake()
							self.exportFBX()
				else :
					cmds.text('detection',label='Excel�����뵱ǰ������ƥ��',ebg=1,e=1)
					break

	def bake(self,*args):
		
		self.obj=cmds.ls(sl=1)

		#����ѡ���������ò�ɾ��ѡ����������ƿռ�
		path_reference=cmds.referenceQuery(self.obj[0], f=True )
		cmds.file(path_reference, ir=1)
		cmds.parent(self.obj,world=True)#�����
		self.obj_namespace=str(self.obj[0]).split(':',1)[0]+':'
		cmds.namespace(removeNamespace = self.obj_namespace, mergeNamespaceWithParent = True)
		self.obj_new=cmds.ls(sl=1)
		
		#ѡ����Ҫ�決�Ķ���
		cmds.select(self.obj_new,hi=1)
		cmds.select(cmds.ls(type= "blendShape"),add=1)
		
		#�決
		self.obj_bake_l=[]
		self.obj_list=cmds.ls(sl=1)
		i=0
		while i<len(self.obj_list):
			self.obj_bake_l.append(str(self.obj_list[i]))
			i+=1
		self.obj_bake_a=str(self.obj_bake_l)[2:-2]
		self.obj_bake=self.obj_bake_a.replace("'",'"')
		#self.start=int(cmds.textFieldGrp('start',text=1,q=1))
		#self.end=int(cmds.textFieldGrp('end',text=1,q=1))
		mel.eval('''bakeResults -simulation true -t "%s:%s" -sampleBy 1 -oversamplingRate 1 -disableImplicitControl true -preserveOutsideKeys true -sparseAnimCurveBake false -removeBakedAttributeFromLayer false -removeBakedAnimFromLayer false -bakeOnOverrideLayer false -minimizeRotation true -controlPoints false -shape true {"%s"}'''%(self.start,self.end,self.obj_bake))
		
		#ɾ������ڵ�
		#cmds.select( '*lower*ipBC*','*upper*ipBC*','*lower*ipBC*_*' )
		#cmds.delete()
		
		#ѡ��������Ҫ�����Ķ���
		cmds.select(self.obj_new,hi=1)
	
	
		
	def exportFBX(self,*args):
		
		#����FBX�ļ���
		text_path=str(cmds.file(q=1,sn=1)).rsplit('_',1)[0]+'_FBX'
		if 'Mb' in text_path:
			text_path_old=text_path.replace('Mb','FBX',1)
		elif 'Lighting/File' in text_path:
			text_path_old=text_path.replace('Lighting/File','Animation/FBX',1)#����·��
		else :
			text_path_old=text_path
		save_file_name='/'+str(cmds.file(q=1,sn=1)).rsplit('_',1)[0].rsplit('/',1)[-1]+'_an_'+self.space_name#�����ļ���
		self.export_path=text_path_old+save_file_name
		print(self.export_path)
		#����fbx
		cmds.select(self.obj_new)
		cmds.FBXResetExport()
		mel.eval('FBXExportFileVersion -v FBX201300')
		mel.eval('FBXExportSmoothingGroups -v true')
		mel.eval('FBXExportSmoothMesh -v true')
		#������ӦFBX�ļ���
		if not os.path.exists(self.export_path.rsplit('/',1)[0]):
			os.makedirs(self.export_path.rsplit('/',1)[0])
		
		mel.eval('FBXExport -f "%s" -s'%(self.export_path))
		
		#����������
		collect_g='%s_g'%self.obj_namespace.split(':',1)[0]
		cmds.group(em=True,n=collect_g)
		cmds.parent(self.obj_new,collect_g)

	def openFile(self,*args):
		#��ʾ�����
		cmds.text('detection',label='',align='left',ebg=0,bgc=[1,0,0],e=1)
		
		singleFilter="Ŀ¼"
		self.open_path=str(cmds.fileDialog2(fileFilter=singleFilter,startingDirectory =str(cmds.file(q=1,sn=1)).rsplit('/',1)[0],fm=3)[0])
		cmds.textField('export',text=self.open_path,e=1)
		
	def openExcel(self,*args):
		#��ʾ�����
		cmds.text('detection',label='',align='left',ebg=0,bgc=[1,0,0],e=1)
		
		singleFilter="*.xlsx"
		self.open_path=str(cmds.fileDialog2(fileFilter=singleFilter,startingDirectory=str(cmds.file(q=1,sn=1)).rsplit('/',1)[0],fileMode=4)[0])
		cmds.textField('excel',text=self.open_path,e=1)
	
	def detectionFile(self,*args):
		
		self.excelGet()
		#��ȡ�ļ�·��
		dir_path =cmds.textField('export',text=1,q=1)
		ma_file=[]
		for dirpath, dirnames, filenames in os.walk(dir_path):
			for filename in filenames:
				if '.ma' in filename or '.mb' in filename:
					ma_file.append(filename.rsplit('_',1)[0])
		for ma_name in ma_file:	
			
			if ma_name not in self.start_endK_d:
				cmds.text('detection',label='Excel�뵱ǰ��ѡ�ļ������ݲ�ƥ��',ebg=1,bgc=[1,0,0],e=1)
				print(ma_name,self.start_endK_d)
				break
			else:
				cmds.text('detection',label='Excel�뵱ǰ��ѡ�ļ�������ƥ��',ebg=0,bgc=[0,0,0],e=1)
					
def start():
	win()
	cmds.showWindow()
	
	
	
if __name__=='__main__':
	start()
 


